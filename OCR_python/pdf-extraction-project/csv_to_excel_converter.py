#!/usr/bin/env python3
"""
CSV to Excel - Key-Value Format
Each row becomes Key (Category + Field) and Value.
Section headers are preserved.
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def csv_to_excel_key_value(csv_file: str, excel_file: str):
    csv_path = Path(csv_file)
    excel_path = Path(excel_file)

    if not csv_path.exists():
        print(f"❌ CSV file not found: {csv_file}")
        return

    # Read CSV
    df = pd.read_csv(csv_path, delimiter=",", quotechar='"', keep_default_na=False)

    # Create workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "PDF Content"

    # Styles
    key_font = Font(bold=True)
    section_font = Font(bold=True, size=12)
    wrap_align = Alignment(wrap_text=True, vertical="top")
    section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    row_idx = 1
    for idx, row in df.iterrows():
        category = str(row['Category']).strip()
        field = str(row['Field']).strip()
        value = str(row['Value']).strip()

        # Detect section header
        if category.startswith("==="):
            ws.cell(row=row_idx, column=1, value=category).font = section_font
            ws.cell(row=row_idx, column=1).fill = section_fill
            ws.cell(row=row_idx, column=1).alignment = wrap_align
            row_idx += 1
            continue

        # Combine Category and Field as Key
        key = category
        if field:
            key += f" | {field}"

        ws.cell(row=row_idx, column=1, value=key).font = key_font
        ws.cell(row=row_idx, column=1).alignment = wrap_align
        ws.cell(row=row_idx, column=1).border = thin_border

        ws.cell(row=row_idx, column=2, value=value).alignment = wrap_align
        ws.cell(row=row_idx, column=2).border = thin_border

        row_idx += 1

    # Auto-adjust column widths
    for col_idx in [1, 2]:
        max_length = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 5, 120)

    # Save workbook
    wb.save(excel_path)
    print(f"✅ Excel created successfully: {excel_path}")

if __name__ == "__main__":
    csv_path = "output/pg1_data.csv"
    excel_path = "output/pg1_data_key_value.xlsx"
    csv_to_excel_key_value(csv_path, excel_path)
