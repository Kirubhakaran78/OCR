#!/usr/bin/env python3
"""
Verbose extractor: find the FULL CONTENT section, try to parse a table,
and ALWAYS write an Excel file (either parsed table or raw full-content text).
Run:
  python extract_fullcontent_to_excel_verbose.py C:/path/to/pg1_data.csv
"""
from pathlib import Path
import re
import io
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

FULL_CONTENT_KEYS = [
    "COMPLETE EXTRACTED CONTENT",
    "FULL CONTENT",
    "FULL_CONTENT",
    "FULL_CONTENTS",
    "COMPLETE EXTRACTED CONTENTS",
    "COMPLETE CONTENT",
    "FULL_CONTENT_SECTION",
    "full_content"
]

OUTPUT_XLSX_SUFFIX = "_fullcontent.xlsx"
OUTPUT_CSV_SUFFIX = "_fullcontent_table.csv"

def read_raw_text(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="ignore")

def detect_section_headers(lines):
    headers = []
    for i, ln in enumerate(lines):
        s = ln.strip()
        if not s:
            continue
        for key in FULL_CONTENT_KEYS:
            if s.lower() == key.lower():
                headers.append((i, key))
        if re.fullmatch(r"[A-Z0-9 \-_/&]{2,120}", s) and any(ch.isalpha() for ch in s):
            if re.search(r"\b(CONTENT|COMPLETE|PAGE|EXTRACTED|FULL)\b", s):
                headers.append((i, s))
    return headers

def extract_section(lines, header_idx):
    n = len(lines)
    next_idx = n
    for j in range(header_idx + 1, n):
        s = lines[j].strip()
        if s:
            if re.fullmatch(r"[A-Z0-9 \-_/&]{2,120}", s) and re.search(r"\b(METADATA|SUMMARY|WELL|STANDARDS|SAMPLES|PAGE CONTENT|COMPLETE|FULL|SETTINGS)\b", s):
                next_idx = j
                break
    section_lines = lines[header_idx + 1: next_idx]
    while section_lines and not section_lines[0].strip():
        section_lines.pop(0)
    while section_lines and not section_lines[-1].strip():
        section_lines.pop()
    return section_lines

def find_full_content_section(raw_text: str):
    lines = raw_text.splitlines()
    for i, ln in enumerate(lines):
        if ln.strip().lower() in [k.lower() for k in FULL_CONTENT_KEYS]:
            return extract_section(lines, i), i, ln
    detected = detect_section_headers(lines)
    if detected:
        for idx, txt in detected:
            if re.search(r"\b(COMPLETE|EXTRACTED|FULL)\b", txt, re.I) and re.search(r"\b(CONTENT|CONTENTs|PAGE)\b", txt, re.I):
                return extract_section(lines, idx), idx, txt
        idx, txt = detected[0]
        return extract_section(lines, idx), idx, txt
    return None, None, None

def extract_markdown_pipe_table(lines):
    start = None
    end = None
    for i in range(len(lines)):
        if lines[i].strip().startswith("|") and "|" in lines[i].strip()[1:]:
            if i + 1 < len(lines) and re.search(r"^\s*\|?\s*:?-{2,}\s*(\|.+)?$", lines[i + 1]):
                start = i
                j = i + 2
                while j < len(lines) and lines[j].strip().startswith("|"):
                    j += 1
                end = j
                break
    if start is not None:
        return "\n".join(lines[start:end])
    return None

def extract_csv_like_block(lines):
    block = []
    consec = 0
    blocks = []
    for ln in lines:
        if ln.count(",") >= 1:
            consec += 1
            block.append(ln)
        else:
            if consec >= 2:
                blocks.append(block[:])
            block = []
            consec = 0
    if consec >= 2:
        blocks.append(block)
    if blocks:
        largest = max(blocks, key=len)
        return "\n".join(largest)
    return None

def extract_key_value_block(lines):
    kv_lines = [ln for ln in lines if ":" in ln and re.match(r"^\s*[\w \-()\/]+:\s*", ln)]
    if len(kv_lines) >= 2:
        return kv_lines
    return None

def parse_table_from_section(section_lines):
    if not section_lines:
        return None, "empty"
    md = extract_markdown_pipe_table(section_lines)
    if md:
        try:
            df = pd.read_csv(io.StringIO(md), sep="|", engine="python", skipinitialspace=True)
            df = df.loc[:, [c for c in df.columns if not str(c).startswith("Unnamed") or df[c].notna().any()]]
            df.columns = [str(c).strip() for c in df.columns]
            df = df.dropna(how="all")
            return df, "markdown_pipe"
        except Exception:
            rows = [ [cell.strip() for cell in re.split(r"\s*\|\s*", r.strip("| \t")) if cell.strip()!=""] for r in md.splitlines()]
            if len(rows) >= 2:
                header = rows[0]
                data_rows = rows[2:] if re.search(r"^-{2,}", rows[1][0]) or re.search(r"^-{2,}", "|".join(rows[1])) else rows[1:]
                df = pd.DataFrame(data_rows, columns=header[:len(data_rows[0])])
                return df, "markdown_pipe_manual"
    csv_block = extract_csv_like_block(section_lines)
    if csv_block:
        try:
            df = pd.read_csv(io.StringIO(csv_block), engine="python")
            return df, "csv_block"
        except Exception:
            try:
                df = pd.read_csv(io.StringIO(csv_block), sep=",", engine="python", quotechar='"', skipinitialspace=True)
                return df, "csv_block_relaxed"
            except Exception:
                pass
    kv = extract_key_value_block(section_lines)
    if kv:
        rows = []
        for ln in kv:
            parts = ln.split(":", 1)
            key = parts[0].strip()
            val = parts[1].strip()
            rows.append((key, val))
        df = pd.DataFrame(rows, columns=["Field", "Value"])
        return df, "key_value"
    text = "\n".join(section_lines).strip()
    paras = [p.strip() for p in re.split(r"\n\s*\n", text) if p.strip()]
    if len(paras) > 1:
        return pd.DataFrame({"Paragraph": paras}), "paragraphs"
    return pd.DataFrame({"FullContent": [text]}), "raw_single"

def auto_format_excel(excel_path: Path):
    wb = load_workbook(excel_path)
    for ws in wb.worksheets:
        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    cell.alignment = Alignment(wrap_text=True)
                    l = len(str(cell.value))
                    if l > max_len:
                        max_len = l
            ws.column_dimensions[col_letter].width = min(max_len + 2, 80)
        ws.freeze_panes = "A2"
    wb.save(excel_path)

def main(csv_path: Path):
    print("CSV path:", csv_path)
    if not csv_path.exists():
        print("CSV file not found:", csv_path)
        return 2
    raw = read_raw_text(csv_path)
    section_lines, header_idx, header_text = find_full_content_section(raw)
    if section_lines is None:
        print("Full content section not found. Detected header:", header_text, "at index", header_idx)
        print("As a fallback, writing entire CSV text into Excel 'FullContent_Raw' sheet.")
        # fallback: write entire raw text into excel
        out_xlsx = csv_path.with_name(csv_path.stem + OUTPUT_XLSX_SUFFIX)
        df_raw = pd.DataFrame({"FullContent_Raw": [raw[:100000]]})  # limit length for single cell
        with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
            df_raw.to_excel(writer, index=False, sheet_name="FullContent_Raw")
        auto_format_excel(out_xlsx)
        print("✅ Wrote fallback Excel:", out_xlsx)
        return 0
    print(f"Detected full-content header at line {header_idx}: {header_text!r}")
    preview = "\n".join(section_lines[:200])
    print("---- section preview (first 200 lines) ----")
    print(preview[:2000])
    print("---- end preview ----")
    df, method = parse_table_from_section(section_lines)
    print("Parsing method result:", method)
    out_xlsx = csv_path.with_name(csv_path.stem + OUTPUT_XLSX_SUFFIX)
    out_csv  = csv_path.with_name(csv_path.stem + OUTPUT_CSV_SUFFIX)
    if df is None or df.empty:
        print("No tabular data parsed; writing raw section into Excel as 'FullContent_Raw'.")
        df_raw = pd.DataFrame({"FullContent_Raw": ["\n".join(section_lines)]})
        with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
            df_raw.to_excel(writer, index=False, sheet_name="FullContent_Raw")
        auto_format_excel(out_xlsx)
        print("✅ Wrote raw full-content Excel:", out_xlsx)
        return 0
    # Save parsed table
    df.to_csv(out_csv, index=False)
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="FullContent")
    auto_format_excel(out_xlsx)
    print("✅ Parsed table saved as CSV:", out_csv)
    print("✅ Formatted Excel saved as:", out_xlsx)
    return 0

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python extract_fullcontent_to_excel_verbose.py /path/to/pg1_data.csv")
        sys.exit(1)
    csv_path = Path(sys.argv[1])
    sys.exit(main(csv_path))
